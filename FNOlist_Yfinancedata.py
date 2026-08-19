# -*- coding: utf-8 -*-
"""
Ultimate Nifty 50 Analyzer with Multiple Technical Indicators (Optimized)

Key optimizations vs the original:
  1. Single batched yf.download() call for all tickers (threads=True) instead of
     one network round-trip per ticker -> the biggest speed win by far.
  2. WMA and LSMA rolling calculations rewritten as closed-form vectorized numpy
     (sliding_window_view) instead of .rolling().apply(..., raw=True), which is
     an interpreted Python-level loop under the hood. This is 10-50x faster and
     is what matters most when you scale this from Nifty50 to the full NSE
     universe (~1,900-2,100 tickers).
  3. (GEO_MEAN indicator removed entirely per user request — no longer computed.)
  4. Gann time-cycle and Gann target/reversal-zone logic vectorized with numpy
     boolean masks instead of row-by-row Python for-loops.
  5. Stochastic divergence detection vectorized with rolling min/max instead of
     a per-row Python for-loop.
  6. Added an `interval` parameter (default "1d") that is now recorded in an
     "Interval" column in the output CSV, and included in the output filename.
  7. Gann target logic now computes BOTH sides:
       - Gann_Resistance   : upside / resistance target (ceil(sqrt(price))^2)
       - Gann_Support  : downside / support target  (floor(sqrt(price))^2)
     Gann_Reversal_Zone now flags "Resistance" OR "Support" depending on which
     level price is currently closest to, and the composite signal gives a
     +1 bullish nudge when price sits in a Gann support zone (mirroring the
     existing -1 for resistance).

  ── NEW ──────────────────────────────────────────────────────────────────────
  8. Close_Color column:
       "Above_Target"  → today's Close exceeded yesterday's Gann_Resistance
                         (price broke above the Gann resistance level)
       "Below_Support" → today's Close fell below yesterday's Gann_Support
                         (price broke below the Gann support level)
       "Neutral"       → price is between the two Gann levels

  9. Excel (.xlsx) output with actual GREEN / RED cell colour on the CLOSE
     column so you can see the signal at a glance without reading text:
       🟢 Dark green  →  Above_Target  (bullish breakout)
       🔴 Dark red    →  Below_Support (bearish breakdown)
       ⚪ No fill     →  Neutral

  10. NEW: "Stock Name" cells for a specific watchlist of tickers are now
      highlighted SKY BLUE in the Excel output, so they stand out visually
      regardless of their Final_Signal / Close_Color.
"""
import yfinance as yf
import pandas as pd
import numpy as np
from numpy.lib.stride_tricks import sliding_window_view
from datetime import datetime


# ── openpyxl is only needed for the Excel export ──────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("⚠  openpyxl not found — Excel export disabled.  "
          "Run:  pip install openpyxl")

# ── FnO underlyings — the single ticker list this script fetches ────────────
FnO = [
    "^INDIAVIX","^BSESN","NIFTY_TOTAL_MKT.NS", "^NSEBANK", "^NSEI",
        "360ONE.NS", "ABB.NS", "ABCAPITAL.NS", "ADANIENSOL.NS", "ADANIENT.NS", "ADANIGREEN.NS",
    "ADANIPORTS.NS", "ADANIPOWER.NS", "ALKEM.NS", "AMBER.NS", "AMBUJACEM.NS", "ANGELONE.NS",
    "APLAPOLLO.NS", "APOLLOHOSP.NS", "ASHOKLEY.NS", "ASIANPAINT.NS", "ASTRAL.NS", "AUBANK.NS",
    "AUROPHARMA.NS", "AXISBANK.NS", "BAJAJ-AUTO.NS", "BAJAJFINSV.NS", "BAJAJHLDNG.NS",
    "BAJFINANCE.NS", "BANDHANBNK.NS", "BANKBARODA.NS", "BANKINDIA.NS", "BDL.NS", "BEL.NS",
    "BHARATFORG.NS", "BHARTIARTL.NS", "BHEL.NS", "BIOCON.NS", "BLUESTARCO.NS", "BOSCHLTD.NS",
    "BPCL.NS", "BRITANNIA.NS", "BSE.NS", "CAMS.NS", "CANBK.NS", "CDSL.NS", "CGPOWER.NS",
    "CHOLAFIN.NS", "CIPLA.NS", "COALINDIA.NS", "COCHINSHIP.NS", "COFORGE.NS", "COLPAL.NS",
    "CONCOR.NS", "CROMPTON.NS", "CUMMINSIND.NS", "DABUR.NS", "DALBHARAT.NS", "DELHIVERY.NS",
    "DIVISLAB.NS", "DIXON.NS", "DLF.NS", "DMART.NS", "DRREDDY.NS", "EICHERMOT.NS", "ETERNAL.NS",
    "FEDERALBNK.NS", "FORCEMOT.NS", "FORTIS.NS", "GAIL.NS", "GLENMARK.NS", "GMRAIRPORT.NS",
    "GODFRYPHLP.NS", "GODREJCP.NS", "GODREJPROP.NS", "GRASIM.NS", "GVT&D.NS", "HAL.NS",
    "HAVELLS.NS", "HCLTECH.NS", "HDFCAMC.NS", "HDFCBANK.NS", "HDFCLIFE.NS", "HEROMOTOCO.NS",
    "HINDALCO.NS", "HINDPETRO.NS", "HINDUNILVR.NS", "HINDZINC.NS", "HYUNDAI.NS",
    "ICICIBANK.NS", "ICICIGI.NS", "ICICIPRULI.NS", "IDEA.NS", "IDFCFIRSTB.NS", "IEX.NS",
    "INDHOTEL.NS", "INDIANB.NS", "INDIGO.NS", "INDUSINDBK.NS", "INDUSTOWER.NS", "INFY.NS",
    "INOXWIND.NS", "IOC.NS", "IREDA.NS", "IRFC.NS", "ITC.NS", "JINDALSTEL.NS", "JIOFIN.NS",
    "JSWENERGY.NS", "JSWSTEEL.NS", "JUBLFOOD.NS", "KALYANKJIL.NS", "KAYNES.NS", "KEI.NS",
    "KFINTECH.NS", "KOTAKBANK.NS", "KPITTECH.NS", "LAURUSLABS.NS", "LICHSGFIN.NS", "LICI.NS",
    "LODHA.NS", "LT.NS", "LTF.NS", "LTM.NS", "LUPIN.NS", "M&M.NS", "MANAPPURAM.NS", "MANKIND.NS",
    "MARICO.NS", "MARUTI.NS", "MAXHEALTH.NS", "MAZDOCK.NS", "MCX.NS", "MFSL.NS", "MOTHERSON.NS",
    "MOTILALOFS.NS", "MPHASIS.NS", "MUTHOOTFIN.NS", "NAM-INDIA.NS", "NATIONALUM.NS",
    "NAUKRI.NS", "NBCC.NS", "NESTLEIND.NS", "NHPC.NS", "NMDC.NS", "NTPC.NS", "NYKAA.NS",
    "OBEROIRLTY.NS", "OFSS.NS", "OIL.NS", "ONGC.NS", "PAGEIND.NS", "PATANJALI.NS", "PAYTM.NS",
    "PERSISTENT.NS", "PETRONET.NS", "PFC.NS", "PGEL.NS", "PHOENIXLTD.NS", "PIDILITIND.NS",
    "PIIND.NS", "PNB.NS", "PNBHOUSING.NS", "POLICYBZR.NS", "POLYCAB.NS", "POWERGRID.NS",
    "POWERINDIA.NS", "PREMIERENE.NS", "PRESTIGE.NS", "RADICO.NS", "RBLBANK.NS", "RECLTD.NS",
    "RELIANCE.NS", "RVNL.NS", "SAIL.NS", "SBICARD.NS", "SBILIFE.NS", "SBIN.NS", "SHREECEM.NS",
    "SHRIRAMFIN.NS", "SIEMENS.NS", "SOLARINDS.NS", "SONACOMS.NS", "SRF.NS", "SUNPHARMA.NS",
    "SUPREMEIND.NS", "SUZLON.NS", "SWIGGY.NS", "TATACONSUM.NS", "TATAELXSI.NS",
    "TATAPOWER.NS", "TATASTEEL.NS", "TCS.NS", "TECHM.NS", "TIINDIA.NS", "TITAN.NS", "TMPV.NS",
    "TORNTPHARM.NS", "TRENT.NS", "TVSMOTOR.NS", "ULTRACEMCO.NS", "UNIONBANK.NS",
    "UNITDSPR.NS", "UNOMINDA.NS", "UPL.NS", "VBL.NS", "VEDL.NS", "VMM.NS", "VOLTAS.NS",
    "WAAREEENER.NS", "WIPRO.NS", "YESBANK.NS", "ZYDUSLIFE.NS"]

NIFTY50_TICKERS = FnO

# Kept as a single-entry registry (rather than removed outright) purely so
# dashboard.py's existing "Ticker list" dropdown and filename-prefix logic
# keep working unmodified — with one list there's just one dropdown option.
_TICKER_LIST_REGISTRY = {
    "FnO": FnO,
}


def get_active_list_name(tickers) -> str:
    """Return the variable name (from _TICKER_LIST_REGISTRY) whose list
    object matches `tickers` (identity check). Falls back to 'TICKERS' if
    no match is found (e.g. a custom/inline list)."""
    for name, lst in _TICKER_LIST_REGISTRY.items():
        if lst is tickers:
            return name
    return "TICKERS"

# ── Watchlist of tickers to highlight SKY BLUE in the "Stock Name" column
SKYBLUE_TICKERS = {
    "EICHERMOT.NS", "HEROMOTOCO.NS", "SHRIRAMFIN.NS", "TATACONSUM.NS", "INDUSINDBK.NS",
    "JSWSTEEL.NS", "BPCL.NS", "DIVISLAB.NS", "APOLLOHOSP.NS", "NESTLEIND.NS", "BAJAJ-AUTO.NS",
    "CIPLA.NS", "WIPRO.NS", "ADANIENT.NS", "ONGC.NS", "COALINDIA.NS", "HINDALCO.NS", "GRASIM.NS",
    "DRREDDY.NS", "BRITANNIA.NS", "SBILIFE.NS", "HDFCLIFE.NS",
}


# ── Indicator Functions ───────────────────────────────────────────────────────

def calculate_rsi(series, window=14):
    delta = series.diff()
    gain = delta.where(delta > 0, 0).rolling(window).mean()
    loss = -delta.where(delta < 0, 0).rolling(window).mean()
    rs = gain / loss
    return 100 - (100 / (1 + rs))


def calculate_stochastic(df, k_window=12, d_window=5):
    low_min  = df['LOW'].rolling(k_window).min()
    high_max = df['HIGH'].rolling(k_window).max()
    k = 100 * ((df['CLOSE'] - low_min) / (high_max - low_min))
    d = k.rolling(d_window).mean()
    return k.rename('Stoch_K'), d.rename('Stoch_D')


def calculate_bollinger_bands(series, window=20, std_dev=2):
    middle = series.rolling(window).mean()
    std    = series.rolling(window).std()
    upper  = middle + std_dev * std
    lower  = middle - std_dev * std
    width  = ((upper - lower) / middle) * 100
    return middle, upper, lower, width


def bollinger_position(close, upper, lower):
    pos = pd.Series("Inside", index=close.index, dtype=str)
    pos[close > upper] = "Upper_Breakout"
    pos[close < lower] = "Lower_Breakout"
    return pos


def volume_analysis(df):
    df['Volume_SMA']   = df['VOLUME'].rolling(20).mean()
    df['Volume_Ratio'] = (df['VOLUME'] / df['Volume_SMA']).round(2)
    df['OBV']          = (np.sign(df['CLOSE'].diff()) * df['VOLUME']).fillna(0).cumsum()
    df['Volume_Trend'] = np.where(df['VOLUME'] > df['Volume_SMA'], "Rising", "Falling")
    df['Volume_Signal'] = ""
    strong_vol = (df['Volume_Ratio'] > 1.8)
    df.loc[strong_vol & (df['CLOSE'] > df['OPEN']), 'Volume_Signal'] = "Strong_Buy_Vol"
    df.loc[strong_vol & (df['CLOSE'] < df['OPEN']), 'Volume_Signal'] = "Strong_Sell_Vol"
    return df


def simple_gann_time_signals(df):
    bar = np.arange(1, len(df) + 1)
    is_cycle = (bar % 21 == 0) | (bar % 90 == 0)
    df['Gann_Time'] = np.where(is_cycle, "Time_Cycle", "")
    return df


def calculate_gann_targets(df, zone_tolerance=0.018):
    close  = df['CLOSE']
    sqrt_p = np.sqrt(close)

    upside   = np.ceil(sqrt_p)  ** 2
    downside = np.floor(sqrt_p) ** 2

    df['Gann_Resistance']  = upside.round(2)
    df['Gann_Support'] = downside.round(2)

    up_ratio   = (close - upside).abs()   / upside
    down_ratio = (close - downside).abs() / downside

    df['Gann_Reversal_Zone'] = np.select(
        [up_ratio < zone_tolerance, down_ratio < zone_tolerance],
        ['Resistance', 'Support'],
        default=''
    )

    if len(df) > 0:
        df.iloc[0, df.columns.get_loc('Gann_Resistance')]       = np.nan
        df.iloc[0, df.columns.get_loc('Gann_Support')]      = np.nan
        df.iloc[0, df.columns.get_loc('Gann_Reversal_Zone')]= ""
    return df


def detect_stochastic_divergence(df, window=20):
    price = df['CLOSE']
    stoch = df.get('Stoch_D', pd.Series(np.nan, index=df.index))
    span  = window + 1

    roll_min_p = price.rolling(span, min_periods=1).min()
    roll_max_p = price.rolling(span, min_periods=1).max()
    roll_min_s = stoch.rolling(span, min_periods=1).min()
    roll_max_s = stoch.rolling(span, min_periods=1).max()

    bullish = (price <= roll_min_p * 1.001) & (stoch > roll_min_s * 1.02)
    bearish = (price >= roll_max_p * 0.999) & (stoch < roll_max_s * 0.98)

    div   = pd.Series("", index=df.index, dtype=str)
    valid = np.arange(len(df)) >= window
    div[bullish & valid] = "Bullish_Div"
    div[bearish & valid] = "Bearish_Div"
    df['Stoch_Div'] = div
    return df


def calculate_wma(series, window=31):
    weights = np.arange(1, window + 1, dtype=float)
    denom   = weights.sum()
    arr     = series.to_numpy(dtype=float)
    out     = np.full(len(arr), np.nan)
    if len(arr) >= window:
        windows    = sliding_window_view(arr, window)
        out[window - 1:] = windows @ weights / denom
    return pd.Series(out, index=series.index)


def calculate_lsma(series, window=36):
    n   = window
    arr = series.to_numpy(dtype=float)
    out = np.full(len(arr), np.nan)
    if len(arr) >= n:
        x      = np.arange(n, dtype=float)
        sum_x  = x.sum()
        sum_x2 = (x ** 2).sum()
        denom  = n * sum_x2 - sum_x ** 2

        windows  = sliding_window_view(arr, n)
        sum_y    = windows.sum(axis=1)
        sum_xy   = windows @ x

        slope     = (n * sum_xy - sum_x * sum_y) / denom
        intercept = (sum_y - slope * sum_x) / n
        out[n - 1:] = slope * (n - 1) + intercept
    return pd.Series(out, index=series.index)


def calculate_composite_signal(df, buy_threshold=2, sell_threshold=-2):
    # Score is kept purely for REFERENCE / sorting / cell colouring —
    # it no longer decides Final_Signal (see rule-based logic below).
    score = pd.Series(0, index=df.index, dtype=int)

    score += np.select([df['Signal'] == 'LONG', df['Signal'] == 'SHORT'], [1, -1], default=0)
    score += np.select([df['Volume_Signal'] == 'Strong_Buy_Vol', df['Volume_Signal'] == 'Strong_Sell_Vol'], [1, -1], default=0)
    score += np.select([df['Stoch_Div'] == 'Bullish_Div', df['Stoch_Div'] == 'Bearish_Div'], [1, -1], default=0)
    score += np.select([df['BB_Position'] == 'Lower_Breakout', df['BB_Position'] == 'Upper_Breakout'], [1, -1], default=0)
    score += np.select([df['RSI'] < 30, df['RSI'] > 70], [1, -1], default=0)
    score += np.select([df['Diff_Trough'] == 'LONG', df['Diff_Peak'] == 'Short'], [1, -1], default=0)
    score += np.select([df['Gann_Reversal_Zone'] == 'Support', df['Gann_Reversal_Zone'] == 'Resistance'], [1, -1], default=0)

    df['Score'] = score

    # ── Rule-based Final_Signal (replaces the old score-threshold approach) ────
    # BUY  = (trough reversal   OR bullish LSMA-WMA cross)
    #        AND (close broke below Gann support
    #             OR BB lower breakout
    #             OR strong buy volume)
    # SELL = mirror image:
    #        (peak reversal OR bearish LSMA-WMA cross)
    #        AND (close broke above Gann resistance
    #             OR BB upper breakout
    #             OR strong sell volume)
    buy_trend    = (df['Diff_Trough'] == 'LONG') | (df['Signal'] == 'LONG')
    buy_confirm  = (df['Close_Color'] == 'Below_Support') | \
                   (df['BB_Position'] == 'Lower_Breakout') | \
                   (df['Volume_Signal'] == 'Strong_Buy_Vol')

    sell_trend   = (df['Diff_Peak'] == 'Short') | (df['Signal'] == 'SHORT')
    sell_confirm = (df['Close_Color'] == 'Above_Resistance') | \
                   (df['BB_Position'] == 'Upper_Breakout') | \
                   (df['Volume_Signal'] == 'Strong_Sell_Vol')

    df['Final_Signal'] = np.select(
        [buy_trend & buy_confirm, sell_trend & sell_confirm],
        ['BUY', 'SELL'],
        default='NEUTRAL'
    )
    return df


def calculate_signal(lsma, wma):
    above      = lsma > wma
    prev_above = above.shift(1).fillna(False)
    signal     = pd.Series("", index=lsma.index, dtype=str)
    signal[(~prev_above) & above]  = "LONG"
    signal[prev_above   & ~above]  = "SHORT"
    return signal


# ── NEW: Close colour vs previous bar's Gann levels ──────────────────────────
def add_close_color(df: pd.DataFrame) -> pd.DataFrame:
    """
    Colours the Close cell using the CURRENT row's own Gann_Resistance / Gann_Support.

    Note: Gann_Resistance  = ceil(√close)²  is always ≥ Close  }  so Close
          Gann_Support = floor(√close)² is always ≤ Close  }  always sits
          between the two levels.

    The midpoint of the Gann range is therefore used to determine which
    half of the range the price is in:

        midpoint = (Gann_Resistance + Gann_Support) / 2

        Above_Target  : Close > Gann_Resistance (same row)
                        → price exceeded Gann resistance              🔴
        Below_Target  : Close ≤ Gann_Resistance (same row)
                        → price is still below Gann resistance        🟢
    """
    # Close > Gann_Resistance (same row) → RED  (price exceeded Gann resistance)
    # Close < Gann_Resistance (same row) → GREEN (price still below Gann resistance)
    df["Close_Color"] = np.where(
    df["HIGH"] > df["Gann_Resistance"], 
    "Above_Resistance",                    # 🔴 Red
    np.where(
        df["LOW"] < df["Gann_Support"], 
        "Below_Support",                # 🟢 Green
        "Neutral"                      # In between Support & Resistance10
    )

    )
    return df


# ── NEW: Excel export with coloured CLOSE cells ───────────────────────────────
# Colour palette (GitHub-dark inspired, same as the dashboard)
_GREEN_FILL   = None
_RED_FILL     = None
_GREEN_FONT   = None
_RED_FONT     = None
_HDR_FILL     = None
_HDR_FONT     = None
_SKYBLUE_FILL = None
_SKYBLUE_FONT = None

if OPENPYXL_OK:
    _GREEN_FILL   = PatternFill(start_color="1A4731", end_color="1A4731", fill_type="solid")
    _RED_FILL     = PatternFill(start_color="4B1217", end_color="4B1217", fill_type="solid")
    _GREEN_FONT   = Font(color="3FB950", bold=True)
    _RED_FONT     = Font(color="F85149", bold=True)
    _HDR_FILL     = PatternFill(start_color="161B22", end_color="161B22", fill_type="solid")
    _HDR_FONT     = Font(color="8B949E", bold=True)
    # NEW: sky blue highlight for the watchlist tickers (Stock Name column)
    _SKYBLUE_FILL = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    _SKYBLUE_FONT = Font(color="000000", bold=True)


def save_excel_with_colors(df: pd.DataFrame, filepath: str) -> None:
    """
    Saves the DataFrame to an .xlsx file with these enhancements:

    1. CLOSE column coloured per row:
         🔴 dark red    →  Close_Color == "Above_Resistance"  (High > Gann_Resistance)
         🟢 dark green  →  Close_Color == "Below_Support"      (Low  < Gann_Support)

    2. Rows sorted by Final_Signal:  BUY → SELL → NEUTRAL
       (then by Date descending within each group so the freshest bars appear first)

    3. Excel AutoFilter on the header row with the Date column pre-filtered
       to the latest date in the dataset — open the file and the most recent
       candle for every stock is immediately visible without any manual filtering.

    4. NEW: "Stock Name" cell is highlighted SKY BLUE whenever the ticker is
       in SKYBLUE_TICKERS, regardless of its Final_Signal / Close_Color.
    """
    if not OPENPYXL_OK:
        print("  ⚠  openpyxl not available — skipping Excel export.")
        return

    from openpyxl.worksheet.filters import FilterColumn, Filters, DateGroupItem

    # ── 1. Sort: BUY first → SELL → NEUTRAL, then Date descending ────────────
    _signal_order = {"BUY": 0, "SELL": 1, "NEUTRAL": 2}
    df = df.copy()
    df["_sort_key"] = df["Final_Signal"].map(_signal_order).fillna(3)
    df["Date"]      = pd.to_datetime(df["Date"])
    df = df.sort_values(
        ["_sort_key", "Date"],
        ascending=[True, False]
    ).drop("_sort_key", axis=1).reset_index(drop=True)

    latest_date = df["Date"].max()   # used for the date pre-filter below

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Nifty50 Signals"

    cols             = list(df.columns)
    close_col_idx    = cols.index("CLOSE")       + 1   # 1-based for openpyxl
    close_color_idx  = cols.index("Close_Color") + 1   # 1-based
    stock_name_idx   = cols.index("Stock Name")  + 1   # 1-based — NEW
    date_col_0       = cols.index("Date")               # 0-based for FilterColumn

    # ── Header row ────────────────────────────────────────────────────────────
    for c_idx, col_name in enumerate(cols, 1):
        cell           = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill      = _HDR_FILL
        cell.font      = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ─────────────────────────────────────────────────────────────
    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        color_label = row[close_color_idx - 1]   # 0-based tuple index
        stock_name  = row[stock_name_idx - 1]    # 0-based tuple index — NEW

        for c_idx, val in enumerate(row, 1):
            col_name = cols[c_idx - 1]

            # Date column: store as real datetime so Excel date filter works
            # Must strip tzinfo — Excel does not support timezone-aware datetimes
            if col_name == "Date":
                dt        = val.to_pydatetime() if hasattr(val, "to_pydatetime") else val
                write_val = dt.replace(tzinfo=None)
            elif hasattr(val, "isoformat"):
                write_val = str(val)[:19]
            else:
                write_val = val

            cell = ws.cell(row=r_idx, column=c_idx, value=write_val)

            if col_name == "Date":
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

            # Green / red fill on CLOSE cell only
            if c_idx == close_col_idx:
                if color_label == "Above_Resistance":   # High > Gann_Resistance → RED
                    cell.fill = _RED_FILL
                    cell.font = _RED_FONT
                elif color_label == "Below_Support":    # Low < Gann_Support → GREEN
                    cell.fill = _GREEN_FILL
                    cell.font = _GREEN_FONT

            # NEW: Sky blue fill on Stock Name cell for watchlist tickers
            if c_idx == stock_name_idx:
                if stock_name in SKYBLUE_TICKERS:
                    cell.fill = _SKYBLUE_FILL
                    cell.font = _SKYBLUE_FONT

    # ── 2. AutoFilter on the full header range ─────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    # ── 3. Pre-filter Date column to the latest date ──────────────────────────
    fc = FilterColumn(colId=date_col_0)
    fc.filters = Filters(dateGroupItem=[DateGroupItem(
        year=int(latest_date.year),
        month=int(latest_date.month),
        day=int(latest_date.day),
        dateTimeGrouping="day",
    )])
    ws.auto_filter.filterColumn.append(fc)

    # ── Auto-column width + freeze header ─────────────────────────────────────
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 28)

    ws.freeze_panes = "A2"
    wb.save(filepath)


# ── yfinance data source ──────────────────────────────────────────────────
# Every ticker in the FnO list is fetched as "<TICKER>.NS" via yfinance.
# Note: this suffix is only correct for plain NSE equities. Indices
# (NIFTY, BANKNIFTY, SENSEX...), currency pairs (EURINR...), and commodities
# (GOLD, CRUDEOIL...) do NOT actually resolve on Yahoo Finance as
# "<NAME>.NS" — Yahoo uses different conventions for those (e.g. "^NSEI" for
# Nifty, "EURINR=X" for currency pairs, "GC=F" for gold futures). Those
# names will come back empty from yfinance with this suffix; flagging this
# now rather than silently masking it — let me know if you want proper
# Yahoo symbols mapped in for the non-equity names.
def resample_ohlcv(df: pd.DataFrame, rule: str) -> pd.DataFrame:
    """Roll finer OHLCV candles into coarser ones (used for 4h and 1wk,
    neither of which yfinance returns natively for NSE data)."""
    if df.empty:
        return df
    agg = {"OPEN": "first", "HIGH": "max", "LOW": "min", "CLOSE": "last", "VOLUME": "sum"}
    out = df.resample(rule).agg(agg)
    out = out.dropna(subset=["OPEN", "HIGH", "LOW", "CLOSE"], how="all")
    return out


def process_ticker_df(df: pd.DataFrame, ticker: str, requested_interval: str,
                       buy_threshold: int = 2, sell_threshold: int = -2) -> pd.DataFrame:
    """Applies every indicator (unchanged from the original script) to one
    ticker's OHLCV dataframe."""
    df = df.reset_index(drop=True)
    #df["Stock Name"] = ticker.replace(".NS", "").replace(".BO", "").replace("^", "")
    df["Interval"]   = requested_interval
    df["Return"]     = df["CLOSE"].diff().round(2)

    df["WMA"]      = calculate_wma(df["CLOSE"])
    df["LSMA"]     = calculate_lsma(df["CLOSE"])
    df["LSMA-WMA"] = (df["LSMA"] - df["WMA"]).round(2)
    df["Signal"]   = calculate_signal(df["LSMA"], df["WMA"])

    df["RSI"] = calculate_rsi(df["CLOSE"])
    df["Stoch_K"], df["Stoch_D"] = calculate_stochastic(df)

    df["BB_Middle"], df["BB_Upper"], df["BB_Lower"], df["BB_Width"] = \
        calculate_bollinger_bands(df["CLOSE"])
    df["BB_Position"] = bollinger_position(df["CLOSE"], df["BB_Upper"], df["BB_Lower"])

    df = volume_analysis(df)
    df = simple_gann_time_signals(df)
    df = calculate_gann_targets(df)
    df = detect_stochastic_divergence(df)
    df = add_close_color(df)

    resistance_change = df["Gann_Resistance"].diff()
    support_change     = df["Gann_Support"].diff()
    level_shift = pd.Series("", index=df.index, dtype=str)
    level_shift[(resistance_change < 0) | (support_change < 0)] = "Level_Dropped"
    level_shift[support_change > 0] = "Support_Up"
    df["Gann_Level_Shift"] = level_shift

    band_width     = df["Gann_Resistance"] - df["Gann_Support"]
    resistance_pct = (df["Gann_Resistance"] - df["CLOSE"]) / band_width * 100
    support_pct    = (df["CLOSE"] - df["Gann_Support"]) / band_width * 100
    gap_vals = []
    for cc, rp, sp in zip(df["Close_Color"], resistance_pct, support_pct):
        if cc == "Neutral" and pd.notna(rp) and pd.notna(sp):
            gap_vals.append(f"{rp:.2f}% // {sp:.2f}%")
        else:
            gap_vals.append("")
    df["Resistance_Support_Gap"] = gap_vals

    diff = df["LSMA-WMA"]
    df["Diff_Peak"]   = ((diff < diff.shift(1)) & (diff.shift(1) > diff.shift(2))&(diff.shift(2) >  diff.shift(3))&(diff.shift(3) >  diff.shift(4))).map({True: "Short",  False: ""})
    df["Diff_Trough"] = ((diff > diff.shift(1)) & (diff.shift(1) < diff.shift(2))&(diff.shift(2) < diff.shift(3))&(diff.shift(3) >  diff.shift(4))).map({True: "LONG", False: ""})

    df = calculate_composite_signal(df, buy_threshold=buy_threshold, sell_threshold=sell_threshold)
    return df


# ── Main Function ─────────────────────────────────────────────────────────────
def fetch_nifty50_data(n_days: int = 30, interval: str = "1d",
                        buy_threshold: int = 2, sell_threshold: int = -2,
                        ticker_list=None) -> pd.DataFrame:
    """
    ticker_list: optional list of tickers to fetch. Defaults to the FnO list
    (NIFTY50_TICKERS = FnO). Each ticker is fetched from yfinance as
    "<TICKER>.NS" (see the module-level note above on which names this
    suffix is actually correct for).

    interval: yfinance-native intervals ("15m","30m","1h","1d") are passed
    straight through. "4h" and "1wk" are fetched as "1h"/"1d" respectively
    and resampled locally, since yfinance has no native 4h interval for NSE
    data and "1wk" resampling here keeps week boundaries consistent with
    the rest of this script's Friday-close convention.
    """
    tickers = ticker_list if ticker_list is not None else NIFTY50_TICKERS
    requested_interval = interval
    yf_tickers = [t + ".NS" for t in tickers]
    fetch_interval = "1h" if interval == "4h" else ("1d" if interval == "1wk" else interval)

    print(f"Downloading {len(yf_tickers)} tickers in one batched call "
          f"(period={n_days}d, interval={fetch_interval}"
          + (f" -> resampled to {interval}" if interval in ("4h", "1wk") else "") + ") ...")

    raw = yf.download(
        yf_tickers, period=f"{n_days}d", interval=fetch_interval,
        group_by="ticker", auto_adjust=True, threads=True, progress=False,
    )

    if raw.empty:
        raise ValueError("No data fetched. Check connection, tickers, or interval.")

    all_frames = []
    for ticker, yf_ticker in zip(tickers, yf_tickers):
        try:
            if isinstance(raw.columns, pd.MultiIndex):
                if yf_ticker not in raw.columns.get_level_values(0):
                    print(f"  ⚠ No data for {yf_ticker}")
                    continue
                df = raw[yf_ticker].copy()
            else:
                df = raw.copy()

            df = df.dropna(how="all")
            if df.empty:
                print(f"  ⚠ No data for {yf_ticker}")
                continue

            df = df[["Open", "High", "Low", "Close", "Volume"]].copy()
            df.columns = ["OPEN", "HIGH", "LOW", "CLOSE", "VOLUME"]

            if requested_interval in ("4h", "1wk"):
                rule = "4h" if requested_interval == "4h" else "W-FRI"
                df = resample_ohlcv(df, rule)
                if df.empty:
                    print(f"  ⚠ No {requested_interval} candles for {yf_ticker} after resampling")
                    continue

            df = df.reset_index()
            df = df.rename(columns={"Datetime": "Date", "Date": "Date"})

            all_frames.append(process_ticker_df(df, ticker, requested_interval,
                                                  buy_threshold, sell_threshold))
        except Exception as e:
            print(f"  ✗ Error processing {yf_ticker}: {e}")

    if not all_frames:
        raise ValueError("No data fetched. Check connection, tickers, or interval.")

    result = pd.concat(all_frames, ignore_index=True)

    cols = [
        "Stock Name", "Date", "Interval", "Score", "Final_Signal", "Resistance_Support_Gap", "Return",
        "OPEN", "HIGH", "LOW", "CLOSE", "Close_Color", "VOLUME",
        "RSI", "Stoch_K", "Stoch_D", "Stoch_Div",
        "BB_Middle", "BB_Upper", "BB_Lower", "BB_Width", "BB_Position",
        "Volume_SMA", "Volume_Ratio", "OBV", "Volume_Trend", "Volume_Signal",
        "WMA", "LSMA", "LSMA-WMA", "Signal",
        "Gann_Time", "Gann_Resistance", "Gann_Support", "Gann_Reversal_Zone", "Gann_Level_Shift",
        "Diff_Peak", "Diff_Trough",
    ]

    numeric_cols = result.select_dtypes(include=[np.number]).columns
    result[numeric_cols] = result[numeric_cols].round(2)
    result = result[cols]
    result = result.dropna(
        subset=["WMA", "LSMA", "RSI", "BB_Middle", "Volume_SMA"]
    ).reset_index(drop=True)
    return result


# ── Run ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    n        = int(input("Enter number of past days (n): ") or 30)
    interval = input("Enter interval (e.g. 15m, 30m, 1h, 1d, 1wk) [default 1h]: ").strip() or "1h"

    print(f"\nFetching data for all Nifty 50 stocks — past {n} days at '{interval}'...\n")
    data = fetch_nifty50_data(n, interval)

    timestamp     = datetime.now().strftime('%Y%m%d_%H%M')
    safe_interval = interval.replace("/", "-")
    list_name     = get_active_list_name(NIFTY50_TICKERS)
    base_name     = f"{list_name}_{n}days_{safe_interval}_{timestamp}"

    # ── CSV (always saved) ────────────────────────────────────────────────────
    csv_file = base_name + ".csv"
    #data.to_csv(csv_file, index=False)
    #print(f"\n✅  CSV  saved → '{csv_file}'  ({len(data)} rows)")

    # ── Excel with GREEN / RED Close cells + SKY BLUE watchlist tickers ──────
    xlsx_file = base_name + ".xlsx"
    save_excel_with_colors(data, xlsx_file)
    if OPENPYXL_OK:
        green_n = (data["Close_Color"] == "Below_Support").sum()
        red_n   = (data["Close_Color"] == "Above_Resistance").sum()
        print(f"✅  XLSX saved → '{xlsx_file}'")
        print(f"    🟢 Below Gann Support    : {green_n} bars")
        print(f"    🔴 Above Gann Resistance : {red_n} bars")
        print(f"    ⚪ Neutral               : {len(data) - green_n - red_n} bars")
        print(f"    🔵 Sky-blue watchlist tickers highlighted: {sorted(SKYBLUE_TICKERS)}")

    # ── Console summary ───────────────────────────────────────────────────────
    signals = data[(data["Signal"] != "") |
                   (data["Volume_Signal"] != "") | (data["Stoch_Div"] != "")]
    print(f"\n📊 Key signals found: {len(signals)}")
    if not signals.empty:
        print(signals[[
            'Final_Signal', 'Stock Name', 'Date', 'Interval', 'CLOSE', 'Close_Color',
            'Signal', 'Volume_Signal', 'Stoch_Div',
            'BB_Position', 'RSI', 'Score'
        ]].to_string(index=False))

    print(f"\n🧮 Final_Signal breakdown:\n{data['Final_Signal'].value_counts().to_string()}")

    strong_calls = data[data["Final_Signal"] != "NEUTRAL"]
    if not strong_calls.empty:
        print(f"\n🔔 Non-neutral composite calls ({len(strong_calls)}):")
        """print(strong_calls[[
            'Stock Name', 'Date', 'CLOSE', 'Close_Color',
            'Gann_Resistance', 'Gann_Support', 'Score', 'Final_Signal'
        ]].to_string(index=False))"""
