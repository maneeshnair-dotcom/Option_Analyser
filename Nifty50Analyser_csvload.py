# -*- coding: utf-8 -*-
"""
Nifty50Analyser_csvload.py
─────────────────────────────────────────────────────────────────────────────
CSV-driven variant of nifty50_analyzer.py.

This has NOTHING to do with the Streamlit dashboard — it's a standalone
sibling script for when you already have OHLC(V) data in a CSV file (from
any charting tool / broker export / etc.) instead of pulling live data from
yfinance.

Same indicator engine, same output columns, same Excel export — the ONLY
difference is where the data comes from:

    nifty50_analyzer.py        : yf.download(tickers, period=..., interval=...)
    Nifty50Analyser_csvload.py : pd.read_csv(filepath)

WHY THIS WORKS UNCHANGED FOR ANY INTERVAL
──────────────────────────────────────────
Every indicator here (WMA, LSMA, RSI, Stochastic, Bollinger, Volume_SMA,
Gann time-cycles, etc.) operates on a fixed number of *bars* (e.g. "last 21
bars", "last 25 bars") — not on calendar time. So it doesn't matter whether
each row in your CSV represents 15 minutes, 30 minutes, 1 hour, 1 day, or
1 month: the exact same code path just runs on however many rows you feed
it. You don't need to tell it the interval for the maths to work — it's
only used as a display label in the output ("Interval" column) and is
auto-detected from the gaps between your CSV's own Date values if you don't
provide one explicitly.

CSV REQUIREMENTS
─────────────────
Required columns (case-insensitive): Date, Open, High, Low, Close
Optional column: Volume — if missing, Volume-based signals (Volume_Trend,
Volume_Signal, Strong_Buy_Vol/Strong_Sell_Vol) simply won't fire (they
degrade to blank/NaN gracefully); everything else still works normally,
since the BUY/SELL rule already treats volume confirmation as just one of
three OR'd conditions (alongside Close_Color and BB_Position).

Any extra columns in your CSV (Bollinger Bands, MACD, Pivot Points, etc.
from whatever charting tool exported it) are simply ignored — this script
recomputes its own indicators from Open/High/Low/Close rather than trusting
pre-computed columns from elsewhere.

USAGE
──────
    from Nifty50Analyser_csvload import analyze_csv

    data = analyze_csv("NIFTY_50_monthly.csv", stock_name="NIFTY_50")
    # data has the exact same columns as nifty50_analyzer.fetch_nifty50_data()

Or run this file directly and follow the prompts (see the __main__ block).
"""
import re
import os
import warnings
from pathlib import Path

import pandas as pd
import numpy as np
from numpy.lib.stride_tricks import sliding_window_view
from datetime import datetime

# ── openpyxl is only needed for the Excel export ──────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("⚠  openpyxl not found — Excel export disabled.  "
          "Run:  pip install openpyxl")

# ── Indicator Functions (identical to nifty50_analyzer.py — bar-count based,
#    so they work unchanged regardless of what timeframe each row represents)

def calculate_rsi(series, window=14):
    delta = series.diff()
    gain = delta.where(delta > 0, 0).rolling(window).mean()
    loss = -delta.where(delta < 0, 0).rolling(window).mean()
    rs = gain / loss
    return 100 - (100 / (1 + rs))


def calculate_stochastic(df, k_window=12, d_window=5):
    low_min  = df['LOW'].rolling(k_window).min()
    high_max = df['HIGH'].rolling(k_window).max()
    k = 100 * ((df['CLOSE'] - low_min) / (high_max - low_min))
    d = k.rolling(d_window).mean()
    return k.rename('Stoch_K'), d.rename('Stoch_D')


def calculate_bollinger_bands(series, window=20, std_dev=2):
    middle = series.rolling(window).mean()
    std    = series.rolling(window).std()
    upper  = middle + std_dev * std
    lower  = middle - std_dev * std
    width  = ((upper - lower) / middle) * 100
    return middle, upper, lower, width


def bollinger_position(close, upper, lower):
    pos = pd.Series("Inside", index=close.index, dtype=str)
    pos[close > upper] = "Upper_Breakout"
    pos[close < lower] = "Lower_Breakout"
    return pos


def volume_analysis(df):
    """Works fine even when VOLUME is entirely NaN (no Volume column in the
    CSV) — every derived column just ends up NaN/blank for that stock and
    the BUY/SELL rule silently ignores that OR-branch."""
    df['Volume_SMA']   = df['VOLUME'].rolling(20).mean()
    df['Volume_Ratio'] = (df['VOLUME'] / df['Volume_SMA']).round(2)
    df['OBV']          = (np.sign(df['CLOSE'].diff()) * df['VOLUME']).fillna(0).cumsum()
    df['Volume_Trend'] = np.where(df['VOLUME'] > df['Volume_SMA'], "Rising", "Falling")
    df['Volume_Signal'] = ""
    strong_vol = (df['Volume_Ratio'] > 1.8)
    df.loc[strong_vol & (df['CLOSE'] > df['OPEN']), 'Volume_Signal'] = "Strong_Buy_Vol"
    df.loc[strong_vol & (df['CLOSE'] < df['OPEN']), 'Volume_Signal'] = "Strong_Sell_Vol"
    return df


def simple_gann_time_signals(df):
    bar = np.arange(1, len(df) + 1)
    is_cycle = (bar % 21 == 0) | (bar % 90 == 0)
    df['Gann_Time'] = np.where(is_cycle, "Time_Cycle", "")
    return df


def calculate_gann_targets(df, zone_tolerance=0.018):
    close  = df['CLOSE']
    sqrt_p = np.sqrt(close)

    upside   = np.ceil(sqrt_p)  ** 2
    downside = np.floor(sqrt_p) ** 2

    df['Gann_Resistance']  = upside.round(2)
    df['Gann_Support'] = downside.round(2)

    up_ratio   = (close - upside).abs()   / upside
    down_ratio = (close - downside).abs() / downside

    df['Gann_Reversal_Zone'] = np.select(
        [up_ratio < zone_tolerance, down_ratio < zone_tolerance],
        ['Resistance', 'Support'],
        default=''
    )

    if len(df) > 0:
        df.iloc[0, df.columns.get_loc('Gann_Resistance')]       = np.nan
        df.iloc[0, df.columns.get_loc('Gann_Support')]      = np.nan
        df.iloc[0, df.columns.get_loc('Gann_Reversal_Zone')]= ""
    return df


def detect_stochastic_divergence(df, window=20):
    price = df['CLOSE']
    stoch = df.get('Stoch_D', pd.Series(np.nan, index=df.index))
    span  = window + 1

    roll_min_p = price.rolling(span, min_periods=1).min()
    roll_max_p = price.rolling(span, min_periods=1).max()
    roll_min_s = stoch.rolling(span, min_periods=1).min()
    roll_max_s = stoch.rolling(span, min_periods=1).max()

    bullish = (price <= roll_min_p * 1.001) & (stoch > roll_min_s * 1.02)
    bearish = (price >= roll_max_p * 0.999) & (stoch < roll_max_s * 0.98)

    div   = pd.Series("", index=df.index, dtype=str)
    valid = np.arange(len(df)) >= window
    div[bullish & valid] = "Bullish_Div"
    div[bearish & valid] = "Bearish_Div"
    df['Stoch_Div'] = div
    return df


def calculate_wma(series, window=21):
    weights = np.arange(1, window + 1, dtype=float)
    denom   = weights.sum()
    arr     = series.to_numpy(dtype=float)
    out     = np.full(len(arr), np.nan)
    if len(arr) >= window:
        windows    = sliding_window_view(arr, window)
        out[window - 1:] = windows @ weights / denom
    return pd.Series(out, index=series.index)


def calculate_lsma(series, window=25):
    n   = window
    arr = series.to_numpy(dtype=float)
    out = np.full(len(arr), np.nan)
    if len(arr) >= n:
        x      = np.arange(n, dtype=float)
        sum_x  = x.sum()
        sum_x2 = (x ** 2).sum()
        denom  = n * sum_x2 - sum_x ** 2

        windows  = sliding_window_view(arr, n)
        sum_y    = windows.sum(axis=1)
        sum_xy   = windows @ x

        slope     = (n * sum_xy - sum_x * sum_y) / denom
        intercept = (sum_y - slope * sum_x) / n
        out[n - 1:] = slope * (n - 1) + intercept
    return pd.Series(out, index=series.index)


def calculate_composite_signal(df, buy_threshold=2, sell_threshold=-2):
    # Score is kept purely for REFERENCE / sorting / cell colouring —
    # it does not decide Final_Signal (see rule-based logic below).
    score = pd.Series(0, index=df.index, dtype=int)

    score += np.select([df['Signal'] == 'LONG', df['Signal'] == 'SHORT'], [1, -1], default=0)
    score += np.select([df['Volume_Signal'] == 'Strong_Buy_Vol', df['Volume_Signal'] == 'Strong_Sell_Vol'], [1, -1], default=0)
    score += np.select([df['Stoch_Div'] == 'Bullish_Div', df['Stoch_Div'] == 'Bearish_Div'], [1, -1], default=0)
    score += np.select([df['BB_Position'] == 'Lower_Breakout', df['BB_Position'] == 'Upper_Breakout'], [1, -1], default=0)
    score += np.select([df['RSI'] < 30, df['RSI'] > 70], [1, -1], default=0)
    score += np.select([df['Diff_Trough'] == 'LONG', df['Diff_Peak'] == 'Short'], [1, -1], default=0)
    score += np.select([df['Gann_Reversal_Zone'] == 'Support', df['Gann_Reversal_Zone'] == 'Resistance'], [1, -1], default=0)

    df['Score'] = score

    # ── Rule-based Final_Signal ──────────────────────────────────────────────
    # BUY  = (trough reversal OR bullish LSMA-WMA cross)
    #        AND (close broke below Gann support OR BB lower breakout
    #             OR strong buy volume)
    # SELL = mirror image using peak reversal / bearish cross / resistance /
    #        upper breakout / strong sell volume.
    buy_trend    = (df['Diff_Trough'] == 'LONG') | (df['Signal'] == 'LONG')
    buy_confirm  = (df['Close_Color'] == 'Below_Support') | \
                   (df['BB_Position'] == 'Lower_Breakout') | \
                   (df['Volume_Signal'] == 'Strong_Buy_Vol')

    sell_trend   = (df['Diff_Peak'] == 'Short') | (df['Signal'] == 'SHORT')
    sell_confirm = (df['Close_Color'] == 'Above_Resistance') | \
                   (df['BB_Position'] == 'Upper_Breakout') | \
                   (df['Volume_Signal'] == 'Strong_Sell_Vol')

    df['Final_Signal'] = np.select(
        [buy_trend & buy_confirm, sell_trend & sell_confirm],
        ['BUY', 'SELL'],
        default='NEUTRAL'
    )
    return df


def calculate_signal(lsma, wma):
    above      = lsma > wma
    prev_above = above.shift(1).fillna(False)
    signal     = pd.Series("", index=lsma.index, dtype=str)
    signal[(~prev_above) & above]  = "LONG"
    signal[prev_above   & ~above]  = "SHORT"
    return signal


def add_close_color(df: pd.DataFrame) -> pd.DataFrame:
    """Colours the Close cell using the CURRENT row's own Gann levels.
    Above_Resistance (High > Gann_Resistance) / Below_Support (Low < Gann_Support) / Neutral."""
    df["Close_Color"] = np.where(
        df["HIGH"] > df["Gann_Resistance"],
        "Above_Resistance",
        np.where(
            df["LOW"] < df["Gann_Support"],
            "Below_Support",
            "Neutral"
        )
    )
    return df


# ── Excel export with coloured CLOSE cells (identical to nifty50_analyzer.py) ─
_GREEN_FILL   = None
_RED_FILL     = None
_GREEN_FONT   = None
_RED_FONT     = None
_HDR_FILL     = None
_HDR_FONT     = None

if OPENPYXL_OK:
    _GREEN_FILL   = PatternFill(start_color="1A4731", end_color="1A4731", fill_type="solid")
    _RED_FILL     = PatternFill(start_color="4B1217", end_color="4B1217", fill_type="solid")
    _GREEN_FONT   = Font(color="3FB950", bold=True)
    _RED_FONT     = Font(color="F85149", bold=True)
    _HDR_FILL     = PatternFill(start_color="161B22", end_color="161B22", fill_type="solid")
    _HDR_FONT     = Font(color="8B949E", bold=True)


def save_excel_with_colors(df: pd.DataFrame, filepath: str) -> None:
    """Same behaviour as nifty50_analyzer.save_excel_with_colors():
    green/red CLOSE cell fill, BUY→SELL→NEUTRAL sort, AutoFilter pre-filtered
    to the latest date."""
    if not OPENPYXL_OK:
        print("  ⚠  openpyxl not available — skipping Excel export.")
        return

    from openpyxl.worksheet.filters import FilterColumn, Filters, DateGroupItem

    _signal_order = {"BUY": 0, "SELL": 1, "NEUTRAL": 2}
    df = df.copy()
    df["_sort_key"] = df["Final_Signal"].map(_signal_order).fillna(3)
    df["Date"]      = pd.to_datetime(df["Date"])
    df = df.sort_values(
        ["_sort_key", "Date"],
        ascending=[True, False]
    ).drop("_sort_key", axis=1).reset_index(drop=True)

    latest_date = df["Date"].max()

    wb = Workbook()
    ws = wb.active
    ws.title = "Signals"

    cols             = list(df.columns)
    close_col_idx    = cols.index("CLOSE")       + 1
    close_color_idx  = cols.index("Close_Color") + 1
    date_col_0       = cols.index("Date")

    for c_idx, col_name in enumerate(cols, 1):
        cell           = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill      = _HDR_FILL
        cell.font      = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        color_label = row[close_color_idx - 1]

        for c_idx, val in enumerate(row, 1):
            col_name = cols[c_idx - 1]

            if col_name == "Date":
                dt        = val.to_pydatetime() if hasattr(val, "to_pydatetime") else val
                write_val = dt.replace(tzinfo=None)
            elif hasattr(val, "isoformat"):
                write_val = str(val)[:19]
            else:
                write_val = val

            cell = ws.cell(row=r_idx, column=c_idx, value=write_val)

            if col_name == "Date":
                cell.number_format = "YYYY-MM-DD HH:MM:SS"

            if c_idx == close_col_idx:
                if color_label == "Above_Resistance":
                    cell.fill = _RED_FILL
                    cell.font = _RED_FONT
                elif color_label == "Below_Support":
                    cell.fill = _GREEN_FILL
                    cell.font = _GREEN_FONT

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    fc = FilterColumn(colId=date_col_0)
    fc.filters = Filters(dateGroupItem=[DateGroupItem(
        year=int(latest_date.year),
        month=int(latest_date.month),
        day=int(latest_date.day),
        dateTimeGrouping="day",
    )])
    ws.auto_filter.filterColumn.append(fc)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 28)

    ws.freeze_panes = "A2"
    wb.save(filepath)


# ── CSV ingestion helpers ──────────────────────────────────────────────────────

def _parse_csv_dates(series: pd.Series) -> pd.Series:
    """
    Robust date parser. Tries a plain pd.to_datetime first (works for most
    clean CSV exports); if that fails (e.g. the messy
    'Mon Apr 01 2013 00:00:00 GMT+0530 (India Standard Time)' format some
    charting tools export), strips the trailing 'GMT...' timezone text and
    retries with an explicit format.
    """
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            return pd.to_datetime(series)
    except Exception:
        pass
    cleaned = series.astype(str).str.replace(r"\s*GMT.*$", "", regex=True)
    try:
        return pd.to_datetime(cleaned, format="%a %b %d %Y %H:%M:%S")
    except Exception:
        # last resort — let pandas guess row by row
        return pd.to_datetime(cleaned, errors="coerce")


def _detect_interval_label(dates: pd.Series) -> str:
    """Auto-detect a friendly interval label ('15m'/'30m'/'1h'/'4h'/'1d'/
    '1wk'/'1mo') from the median gap between consecutive (sorted) dates.
    Only used when the caller doesn't explicitly pass `interval=`."""
    d = dates.sort_values().dropna()
    if len(d) < 2:
        return "unknown"
    median_minutes = d.diff().dropna().dt.total_seconds().median() / 60
    if median_minutes < 20:      return "15m"
    if median_minutes < 45:      return "30m"
    if median_minutes < 180:     return "1h"
    if median_minutes < 1200:    return "4h"
    if median_minutes < 4320:    return "1d"
    if median_minutes < 14400:   return "1wk"
    return "1mo"


def _find_col(df: pd.DataFrame, *candidates) -> str:
    """Case-insensitive column lookup — returns the actual column name in
    df that matches one of the candidate names, or raises a clear error."""
    lower_map = {c.lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower_map:
            return lower_map[cand.lower()]
    raise KeyError(f"None of {candidates} found in CSV columns: {list(df.columns)}")


# ── Standalone CSV loader ────────────────────────────────────────────────────
def load_data(filepath: str) -> pd.DataFrame:
    """
    Load and prepare OHLC(V) data from a CSV export (Date, Open, High, Low,
    Close, optional Volume — any extra indicator columns the charting tool
    tacked on, e.g. PMO/MACD/Bollinger/Pivot columns, are ignored).

    - Columns are matched by NAME (case-insensitive), not position, so it
      doesn't matter whether Volume is missing, is the 6th column, or is
      the last column.
    - Dates are parsed robustly, including the messy
      'Wed Jun 24 2026 15:15:00 GMT+0530 (India Standard Time)' format some
      charting tools export.

    Returns a DataFrame with columns: Date, Open, High, Low, Close, Volume
    (Volume is all-NaN if the CSV didn't have one).
    """
    path = Path(filepath)
    if not path.exists() and path.suffix.lower() != ".csv" and path.with_suffix(".csv").exists():
        path = path.with_suffix(".csv")
    if not path.exists():
        raise FileNotFoundError(
            f"Could not find '{filepath}'.\n"
            f"  Looked for: {path.resolve()}\n"
            f"  Tip: make sure the filename includes '.csv' and that it's in the "
            f"folder you're running the script from (or give the full path)."
        )
    raw = pd.read_csv(path)

    date_col  = _find_col(raw, "Date", "Datetime")
    open_col  = _find_col(raw, "Open")
    high_col  = _find_col(raw, "High")
    low_col   = _find_col(raw, "Low")
    close_col = _find_col(raw, "Close")
    try:
        vol_col = _find_col(raw, "Volume", "Vol")
    except KeyError:
        vol_col = None

    df = pd.DataFrame({
        "Date":  _parse_csv_dates(raw[date_col]),
        "Open":  pd.to_numeric(raw[open_col],  errors="coerce"),
        "High":  pd.to_numeric(raw[high_col],  errors="coerce"),
        "Low":   pd.to_numeric(raw[low_col],   errors="coerce"),
        "Close": pd.to_numeric(raw[close_col], errors="coerce"),
    })
    df["Volume"] = pd.to_numeric(raw[vol_col], errors="coerce") if vol_col else np.nan
    if vol_col is None:
        print(f"  ⚠ No Volume column found in {path.name} — "
              f"volume-based signals will be blank for this file.")

    df = df.dropna(subset=["Date", "Open", "High", "Low", "Close"]).sort_values("Date").reset_index(drop=True)
    if df.empty:
        raise ValueError(f"No usable rows found in {filepath} after parsing.")
    return df


# ── Main Function — single CSV ─────────────────────────────────────────────────
def analyze_csv(filepath: str, stock_name: str = None, interval: str = None,
                 buy_threshold: int = 2, sell_threshold: int = -2) -> pd.DataFrame:
    """
    Load OHLC(V) data from a CSV and run the exact same indicator pipeline
    used by nifty50_analyzer.fetch_nifty50_data() — output has identical
    columns, so it's a drop-in replacement wherever you'd otherwise use that
    function's return value (including feeding save_excel_with_colors()).

    filepath  : path to the CSV file. Must have Date/Open/High/Low/Close
                columns (case-insensitive); Volume is optional.
    stock_name: label for the "Stock Name" column. If not given, it's
                derived from the filename (text before the first run of
                underscores — e.g. "NIFTY_50__2026...csv" -> "NIFTY_50").
    interval  : label for the "Interval" column (e.g. "15m", "1h", "1mo").
                If not given, it's auto-detected from the median gap
                between consecutive dates in the file.
    """
    path = Path(filepath)
    df = load_data(path)
    df = df.rename(columns={"Open": "OPEN", "High": "HIGH", "Low": "LOW",
                             "Close": "CLOSE", "Volume": "VOLUME"})

    if stock_name is None:
        stock_name = re.split(r"_{2,}", path.stem)[0].strip("_") or path.stem
    if interval is None:
        interval = _detect_interval_label(df["Date"])
        print(f"  ℹ Auto-detected interval for {stock_name}: '{interval}' "
              f"(median gap between rows)")

    df["Stock Name"] = stock_name
    df["Interval"]   = interval

    # NEW: Return = current candle CLOSE minus previous candle CLOSE
    df["Return"]   = df["CLOSE"].diff().round(2)

    # Core indicators
    df["WMA"]      = calculate_wma(df["CLOSE"])
    df["LSMA"]     = calculate_lsma(df["CLOSE"])
    df["LSMA-WMA"] = (df["LSMA"] - df["WMA"]).round(2)
    df["Signal"]   = calculate_signal(df["LSMA"], df["WMA"])

    # Additional indicators
    df["RSI"]              = calculate_rsi(df["CLOSE"])
    df["Stoch_K"], df["Stoch_D"] = calculate_stochastic(df)

    # Bollinger Bands
    df["BB_Middle"], df["BB_Upper"], df["BB_Lower"], df["BB_Width"] = \
        calculate_bollinger_bands(df["CLOSE"])
    df["BB_Position"] = bollinger_position(df["CLOSE"], df["BB_Upper"], df["BB_Lower"])

    # Volume
    df = volume_analysis(df)

    # Gann
    df = simple_gann_time_signals(df)
    df = calculate_gann_targets(df)
    df = detect_stochastic_divergence(df)

    # Close colour vs this row's own Gann levels (needed as a mandatory gate
    # for Final_Signal, see calculate_composite_signal)
    df = add_close_color(df)

    # Gann level-shift tracker (this bar's Gann levels vs the previous bar's)
    resistance_change = df["Gann_Resistance"].diff()
    support_change     = df["Gann_Support"].diff()
    level_shift = pd.Series("", index=df.index, dtype=str)
    level_shift[(resistance_change < 0) | (support_change < 0)] = "Level_Dropped"
    level_shift[support_change > 0] = "Support_Up"
    df["Gann_Level_Shift"] = level_shift

    # Resistance/Support gap tracker (only when Close_Color == "Neutral")
    band_width     = df["Gann_Resistance"] - df["Gann_Support"]
    resistance_pct = (df["Gann_Resistance"] - df["CLOSE"]) / band_width * 100
    support_pct    = (df["CLOSE"] - df["Gann_Support"]) / band_width * 100
    gap_vals = []
    for cc, rp, sp in zip(df["Close_Color"], resistance_pct, support_pct):
        if cc == "Neutral" and pd.notna(rp) and pd.notna(sp):
            gap_vals.append(f"{rp:.2f}% // {sp:.2f}%")
        else:
            gap_vals.append("")
    df["Resistance_Support_Gap"] = gap_vals

    # Peak / Trough
    diff = df["LSMA-WMA"]
    df["Diff_Peak"]   = ((diff < diff.shift(1)) & (diff.shift(1) > diff.shift(2))).map({True: "Short",  False: ""})
    df["Diff_Trough"] = ((diff > diff.shift(1)) & (diff.shift(1) < diff.shift(2))).map({True: "LONG", False: ""})

    # Composite signal
    df = calculate_composite_signal(df, buy_threshold=buy_threshold, sell_threshold=sell_threshold)

    cols = [
        "Stock Name", "Date", "Interval", "Score", "Final_Signal", "Resistance_Support_Gap", "Return",
        "OPEN", "HIGH", "LOW", "CLOSE", "Close_Color", "VOLUME",
        "RSI", "Stoch_K", "Stoch_D", "Stoch_Div",
        "BB_Middle", "BB_Upper", "BB_Lower", "BB_Width", "BB_Position",
        "Volume_SMA", "Volume_Ratio", "OBV", "Volume_Trend", "Volume_Signal",
        "WMA", "LSMA", "LSMA-WMA", "Signal",
        "Gann_Time", "Gann_Resistance", "Gann_Support", "Gann_Reversal_Zone", "Gann_Level_Shift",
        "Diff_Peak", "Diff_Trough",
    ]

    numeric_cols = df.select_dtypes(include=[np.number]).columns
    df[numeric_cols] = df[numeric_cols].round(2)
    df = df[cols]
    df = df.dropna(subset=["WMA", "LSMA", "RSI", "BB_Middle"]).reset_index(drop=True)
    return df


# ── Main Function — a folder of CSVs (one per stock) ───────────────────────────
def analyze_csv_folder(folder_path: str, buy_threshold: int = 2,
                        sell_threshold: int = -2) -> pd.DataFrame:
    """
    Runs analyze_csv() on every *.csv file in `folder_path` and concatenates
    the results — mirrors nifty50_analyzer.fetch_nifty50_data()'s multi-
    ticker behaviour for when you have several instruments' CSVs at once.
    Stock names are auto-derived from each filename.
    """
    folder = Path(folder_path)
    csv_files = sorted(folder.glob("*.csv"))
    if not csv_files:
        raise ValueError(f"No .csv files found in {folder_path}")

    frames = []
    for f in csv_files:
        try:
            frames.append(analyze_csv(f, buy_threshold=buy_threshold, sell_threshold=sell_threshold))
        except Exception as e:
            print(f"  ✗ Error processing {f.name}: {e}")

    if not frames:
        raise ValueError("No CSV files could be processed successfully.")
    return pd.concat(frames, ignore_index=True)


# ── Run ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # If your CSV is in the SAME folder you're running this script from, you
    # can just type the filename (e.g. "NIFTY_50.csv") — no need for a full path.
    path_in = input("Enter CSV filename (or folder of CSVs) — same folder as this script works fine: ").strip().strip('"')

    if os.path.isdir(path_in):
        print(f"\nFolder detected — analyzing every CSV in '{path_in}'...\n")
        data = analyze_csv_folder(path_in)
        base_label = Path(path_in).name
    else:
        name_in     = input("Name for this stock/index [blank = auto-detect from filename]: ").strip() or None
        interval_in = input("Interval label, e.g. 15m/30m/1h/1d/1mo [blank = auto-detect]: ").strip() or None
        print(f"\nAnalyzing '{path_in}'...\n")
        data = analyze_csv(path_in, stock_name=name_in, interval=interval_in)
        base_label = data["Stock Name"].iloc[0]

    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    base_name = f"{base_label}_csvload_{timestamp}"

    xlsx_file = base_name + ".xlsx"
    save_excel_with_colors(data, xlsx_file)
    if OPENPYXL_OK:
        green_n = (data["Close_Color"] == "Below_Support").sum()
        red_n   = (data["Close_Color"] == "Above_Resistance").sum()
        print(f"✅  XLSX saved → '{xlsx_file}'")
        print(f"    🟢 Below Gann Support    : {green_n} bars")
        print(f"    🔴 Above Gann Resistance : {red_n} bars")
        print(f"    ⚪ Neutral               : {len(data) - green_n - red_n} bars")

    print(f"\n🧮 Final_Signal breakdown:\n{data['Final_Signal'].value_counts().to_string()}")

    strong_calls = data[data["Final_Signal"] != "NEUTRAL"]
    if not strong_calls.empty:
        print(f"\n🔔 Non-neutral composite calls ({len(strong_calls)}):")
        print(strong_calls[[
            'Stock Name', 'Date', 'Interval', 'CLOSE', 'Close_Color',
            'Gann_Resistance', 'Gann_Support', 'Score', 'Final_Signal'
        ]].to_string(index=False))